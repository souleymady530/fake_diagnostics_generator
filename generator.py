# -*- coding: utf-8 -*-
"""
Created on Tue Dec 13 09:24:07 2022

@author: OFilwende
"""

#import de librairie
#
from openpyxl import load_workbook
from itertools import combinations
from itertools import permutations
import numpy as np



#lecture des données et formatage des differentes liste
wb = load_workbook("A_liste_maladie.xlsx")
sheet = wb.active

dico={}
#la premiere ligne contient les titres des colonnes
for i in range(2,11):
    #on fait une concatenation des champs en separants les champs par des ;
    ligne=str(sheet.cell(i,4).value)+";"+str(sheet.cell(i,5).value)+";"+str(sheet.cell(i,6).value)+";"+str(sheet.cell(i,7).value)+";"+str(sheet.cell(i,8).value+";"+str(sheet.cell(i,9).value)+";"+str(sheet.cell(i,10).value)+";"+
        str(sheet.cell(i,11).value))
    dico[str(sheet.cell(i,3).value)]=tuple(ligne.split(";"))
    

liste_true_diagnostics={}
tab_combinaisons_symptomes=[];
liste_total_symptomes=list()

for key in dico:
    row=np.asarray(dico[key])
    liste_diagnostic_exacte=list();
    for element in list(permutations(row[2:10])):
          liste_diagnostic_exacte.append(element) 
    
    liste_true_diagnostics[key]=liste_diagnostic_exacte
    for i in range(2,8):
        liste_total_symptomes.append(row[i])


#suppression des doublons dans la liste des symptomes
liste_symptomes_unique=list()
for symptomes in liste_total_symptomes:
    if symptomes not in liste_symptomes_unique:
        liste_symptomes_unique.append(symptomes)

#combinaisons des symptomes pour fournir une super liste de diagnostiques (qui est lui aussi une liste de symptomes)
liste_fake_diagnostics=list();
liste_fake_diagnostics=list(combinations(liste_symptomes_unique,6))


#this section combines fake_diagnostics and true. true diagnostics is true 
#even we change order of symptomes in its symptoms list.
all_diagnostics=[]
for liste in liste_true_diagnostics.values():
    for liste2 in liste:
        all_diagnostics.append(liste2);

all_diagnostics+=liste_fake_diagnostics;





fake_diagnostic_matched={}

nbreBingo=0;


for element in all_diagnostics[:50000]:
   # print(element)
    for disease,diagnostic in liste_true_diagnostics.items():
       # print(diagnostic)
        if diagnostic.count(element)>0:
            fake_diagnostic_matched[disease]=element
            nbreBingo+=1
        else:
             fake_diagnostic_matched["INDETERMINE"]=element
           
#decommenter pour connaitre le nombre de bingo sur l interval choisit
#print(nbreBingo)

"""
le dico fake_diagnostic_matched contient la liste generale des diagnostics (faux comme vrai) etiquetté de leur
maladie ("INDETERMINE" si le diagnostics ne correspond a aucun des vraie diagnostics)
"""   
 